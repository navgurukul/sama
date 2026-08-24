#!/usr/bin/env python3
"""Seed sample data into the sheet-first schema from Dev Laptop Data.xlsx.

This script is reproducible for local Postgres and RDS.
It inserts sample rows for base tables from workbook sheets and creates
synthetic rows for control/traceability tables.

Usage:
  python scripts/db/seed_sample_data.py --database-url postgresql://postgres@localhost:5432/sama
  python scripts/db/seed_sample_data.py --database-url "$DATABASE_URL" --schema sama_ops --truncate
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


def _utc_now() -> dt.datetime:
    return dt.datetime.now(dt.UTC)


def _validate_identifier(name: str, label: str) -> str:
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        raise ValueError(f"Invalid {label}: {name!r}")
    return name


def _connect(database_url: str):
    try:
        import psycopg  # type: ignore

        conn = psycopg.connect(database_url)

        def close() -> None:
            conn.close()

        return conn, close
    except Exception as exc:
        raise RuntimeError(
            "psycopg is required. Install with: pip install psycopg[binary]"
        ) from exc


def _load_workbook(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from exc

    return load_workbook(path, data_only=True)


def _row_is_empty(row: Sequence[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _read_sheet_dicts(wb, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if h is None else str(h).strip() for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if _row_is_empty(r):
            continue
        record: Dict[str, Any] = {}
        for i, v in enumerate(r):
            key = headers[i] if i < len(headers) else f"col_{i+1}"
            if key == "":
                key = f"col_{i+1}"
            record[key] = v
        out.append(record)
    return out


def _to_text(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _to_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "")
    try:
        return int(float(s))
    except Exception:
        return None


def _to_numeric(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    s = s.replace("%", "")
    try:
        return float(s)
    except Exception:
        return None


def _try_parse_dt(text: str) -> Optional[dt.datetime]:
    candidates = [
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y",
        "%m/%d/%Y, %H:%M:%S",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
    ]
    for fmt in candidates:
        try:
            parsed = dt.datetime.strptime(text, fmt)
            return parsed
        except Exception:
            continue
    return None


def _to_timestamp(v: Any) -> Optional[dt.datetime]:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime.combine(v, dt.time.min)
    s = str(v).strip()
    if not s:
        return None
    parsed = _try_parse_dt(s)
    return parsed


def _to_date(v: Any) -> Optional[dt.date]:
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if not s:
        return None
    parsed = _try_parse_dt(s)
    return parsed.date() if parsed else None


def _convert(v: Any, kind: str) -> Any:
    if kind == "text":
        return _to_text(v)
    if kind == "int":
        return _to_int(v)
    if kind == "num":
        return _to_numeric(v)
    if kind == "ts":
        return _to_timestamp(v)
    if kind == "date":
        return _to_date(v)
    return v


def _insert_many(conn, schema: str, table: str, cols: List[str], rows: List[Dict[str, Any]], conflict_cols: Optional[List[str]] = None) -> int:
    if not rows:
        return 0
    placeholders = ", ".join(["%s"] * len(cols))
    col_sql = ", ".join(cols)
    conflict_sql = ""
    if conflict_cols:
        conflict_sql = f" ON CONFLICT ({', '.join(conflict_cols)}) DO NOTHING"
    sql = f"INSERT INTO {schema}.{table} ({col_sql}) VALUES ({placeholders}){conflict_sql}"
    values = [tuple(r.get(c) for c in cols) for r in rows]
    with conn.cursor() as cur:
        cur.executemany(sql, values)
    return len(rows)


def _truncate_tables(conn, schema: str) -> None:
    tables = [
        "laptop_checklist_responses",
        "qc_checks",
        "issue_feedback",
        "laptop_stage_runs",
        "laptop_versions",
        "laptop_event_log",
        "laptop_user_map",
        "audit_for_laptops",
        "average_days_count",
        "report",
        "preliminary",
        "pickup",
        "metrics_base",
        "infection",
        "external_registered_ngo",
        "userdetails",
        "laptop_labeling",
    ]
    with conn.cursor() as cur:
        for t in tables:
            cur.execute(f"TRUNCATE TABLE {schema}.{t} RESTART IDENTITY CASCADE")


def seed(database_url: str, schema: str, workbook_path: Path, truncate: bool) -> None:
    schema = _validate_identifier(schema, "schema")
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    conn, close = _connect(database_url)
    wb = _load_workbook(workbook_path)

    try:
        if truncate:
            _truncate_tables(conn, schema)

        inserts_summary: List[Tuple[str, int]] = []

        # 1) laptop_labeling
        laptop_src = _read_sheet_dicts(wb, "Laptop Labeling")
        laptop_rows: List[Dict[str, Any]] = []
        for s in laptop_src:
            raw_cond = _convert(s.get("Condition Status"), "text")
            normalized_cond = None
            if raw_cond:
                cond_upper = str(raw_cond).strip().upper()
                if "GOOD" in cond_upper:
                    normalized_cond = "GOOD"
                elif "BAD" in cond_upper:
                    normalized_cond = "BAD"
                elif "REPAIR" in cond_upper or "NEED" in cond_upper:
                    normalized_cond = "NEEDS_REPAIR"
            
            row = {
                "id": _convert(s.get("ID"), "text"),
                "date_committed": _convert(s.get("Date Committed"), "ts"),
                "donor_company_name": _convert(s.get("Donor Company Name"), "text"),
                "ram": _convert(s.get("RAM"), "text"),
                "rom": _convert(s.get("ROM"), "text"),
                "manufacturer_model": _convert(s.get("Manufacturer Model"), "text"),
                "processor": _convert(s.get("Processor"), "text"),
                "manufacturing_date": _convert(s.get("Manufacturing Date"), "date"),
                "condition_status": normalized_cond,
                "minor_issues": _convert(s.get("Minor Issues"), "text"),
                "major_issues": _convert(s.get("Major Issues"), "text"),
                "other_issues": _convert(s.get("Other Issues"), "text"),
                "inventory_location": _convert(s.get("Inventory Location"), "text"),
                "laptop_weight": _convert(s.get("laptop weight"), "text"),
                "mac_address": _convert(s.get("Mac address"), "text"),
                "status": _convert(s.get("Status"), "text"),
                "working": _convert(s.get("Working"), "text"),
                "battery_capacity": _convert(s.get("Battery Capacity"), "num"),
                "allocated_to": _convert(s.get("Allocated To"), "text"),
                "last_updated_on": _convert(s.get("Last Updated On"), "ts"),
                "last_updated_by": _convert(s.get("Last Updated By"), "text"),
                "assigned_to": _convert(s.get("Assigned To"), "text"),
                "comment_for_issues": _convert(s.get("Comment for the Issues"), "text"),
                "inspection_files": _convert(s.get("Inspection Files"), "text"),
                "activitywatch_pdf": _convert(s.get("ActvityWatch PDF"), "text"),
                "activity_date": _convert(s.get("Date"), "ts"),
                "afk_time": _convert(s.get("AFK Time"), "text"),
                "usage_hours": _convert(s.get("Usage Hours"), "num"),
                "off_times": _convert(s.get("Off Times"), "int"),
                "last_delivery_date": _convert(s.get("Last Delivery Date"), "ts"),
                "refurbishment_date": _convert(s.get("Refurbishment Date"), "ts"),
                "batch": _convert(s.get("Batch"), "text"),
            }
            if row["id"]:
                laptop_rows.append(row)

        laptop_cols = [
            "id", "date_committed", "donor_company_name", "ram", "rom", "manufacturer_model", "processor",
            "manufacturing_date", "condition_status", "minor_issues", "major_issues", "other_issues",
            "inventory_location", "laptop_weight", "mac_address", "status", "working", "battery_capacity",
            "allocated_to", "last_updated_on", "last_updated_by", "assigned_to", "comment_for_issues",
            "inspection_files", "activitywatch_pdf", "activity_date", "afk_time", "usage_hours", "off_times",
            "last_delivery_date", "refurbishment_date", "batch"
        ]
        inserts_summary.append(("laptop_labeling", _insert_many(conn, schema, "laptop_labeling", laptop_cols, laptop_rows, ["id"])))

        with conn.cursor() as cur:
            cur.execute(f"SELECT id FROM {schema}.laptop_labeling")
            laptop_ids = {r[0] for r in cur.fetchall()}

        # 2) userdetails
        user_src = _read_sheet_dicts(wb, "UserDetails")
        user_rows: List[Dict[str, Any]] = []
        for s in user_src:
            row = {
                "id": _convert(s.get("ID"), "int"),
                "ngo": _convert(s.get("Ngo"), "text"),
                "name": _convert(s.get("name"), "text"),
                "email": _convert(s.get("email"), "text"),
                "contact_number": _convert(s.get("contact number"), "text"),
                "address": _convert(s.get("Address"), "text"),
                "address_state": _convert(s.get("Address State"), "text"),
                "id_proof_type": _convert(s.get("ID Proof type"), "text"),
                "id_proof_number": _convert(s.get("ID Proof number"), "text"),
                "qualification": _convert(s.get("Qualification"), "text"),
                "occupation": _convert(s.get("Occupation"), "text"),
                "date_of_birth": _convert(s.get("Date Of Birth"), "date"),
                "use_case": _convert(s.get("Use case"), "text"),
                "family_members_count": _convert(s.get("Number of Family members(who might use the laptop)"), "int"),
                "guardian_occupation": _convert(s.get("Father/Mother/Guardians Occupation"), "text"),
                "family_annual_income": _convert(s.get("Family Annual Income"), "num"),
                "status": _convert(s.get("status"), "text"),
                "laptop_assigned": _convert(s.get("Laptop Assigned"), "text"),
                "id_link": _convert(s.get("ID Link"), "text"),
                "income_certificate_link": _convert(s.get("Income Certificate Link"), "text"),
                "date_time": _convert(s.get("Date-time"), "ts"),
                "doner": _convert(s.get("Doner"), "text"),
            }
            if row["id"] is not None:
                user_rows.append(row)

        user_cols = [
            "id", "ngo", "name", "email", "contact_number", "address", "address_state", "id_proof_type",
            "id_proof_number", "qualification", "occupation", "date_of_birth", "use_case", "family_members_count",
            "guardian_occupation", "family_annual_income", "status", "laptop_assigned", "id_link",
            "income_certificate_link", "date_time", "doner"
        ]
        inserts_summary.append(("userdetails", _insert_many(conn, schema, "userdetails", user_cols, user_rows, ["id"])))

        # 3) pickup
        pickup_src = _read_sheet_dicts(wb, "Pickup")
        pickup_rows: List[Dict[str, Any]] = []
        for s in pickup_src:
            row = {
                "pickup_id": _convert(s.get("Pickup ID"), "text"),
                "donor_company": _convert(s.get("Donor Company"), "text"),
                "poc_name": _convert(s.get("POC Name"), "text"),
                "poc_contact": _convert(s.get("POC Contact"), "text"),
                "poc_email": _convert(s.get("POC Email"), "text"),
                "number_of_laptops": _convert(s.get("Number of Laptops"), "int"),
                "pickup_location": _convert(s.get("Pickup Location"), "text"),
                "pickup_by": _convert(s.get("Pickup By"), "text"),
                "current_date_time": _convert(s.get("Current Date & Time"), "ts"),
                "status": _convert(s.get("Status"), "text"),
                "confirm_pickup_date": _convert(s.get("Confirm Pickup Date"), "ts"),
                "updated_on": _convert(s.get("Updated On"), "ts"),
                "updated_by": _convert(s.get("Updated By"), "text"),
            }
            if row["pickup_id"]:
                pickup_rows.append(row)

        pickup_cols = [
            "pickup_id", "donor_company", "poc_name", "poc_contact", "poc_email", "number_of_laptops",
            "pickup_location", "pickup_by", "current_date_time", "status", "confirm_pickup_date", "updated_on",
            "updated_by"
        ]
        inserts_summary.append(("pickup", _insert_many(conn, schema, "pickup", pickup_cols, pickup_rows, ["pickup_id"])))

        # 4) preliminary
        pre_src = _read_sheet_dicts(wb, "Preliminary")
        pre_rows: List[Dict[str, Any]] = []
        for s in pre_src:
            row = {
                "id": _convert(s.get("Id"), "int"),
                "ngoid": _convert(s.get("NgoId"), "text"),
                "number_of_school": _convert(s.get("Number of school"), "int"),
                "number_of_teacher": _convert(s.get("Number of teacher"), "int"),
                "number_of_student": _convert(s.get("Number of student"), "int"),
                "number_of_female_student": _convert(s.get("Number of Female student"), "int"),
                "states": _convert(s.get("States"), "text"),
                "course": _convert(s.get("Course"), "text"),
                "unit": _convert(s.get("Unit"), "text"),
                "doner": _convert(s.get("Doner"), "text"),
                "request_type": _convert(s.get("requestType"), "text"),
                "ngo_prelim_requests": _convert(s.get("NGOPrelimRequests"), "text"),
            }
            if row["id"] is not None:
                pre_rows.append(row)

        pre_cols = [
            "id", "ngoid", "number_of_school", "number_of_teacher", "number_of_student", "number_of_female_student",
            "states", "course", "unit", "doner", "request_type", "ngo_prelim_requests"
        ]
        inserts_summary.append(("preliminary", _insert_many(conn, schema, "preliminary", pre_cols, pre_rows, ["id"])))

        # 5) report
        rep_src = _read_sheet_dicts(wb, "Report")
        rep_rows: List[Dict[str, Any]] = []
        for s in rep_src:
            row = {
                "id": _convert(s.get("Id"), "int"),
                "ngoid": _convert(s.get("ngoId"), "text"),
                "month": _convert(s.get("Month"), "text"),
                "number_of_teachers_trained": _convert(s.get("Number of Teachers Trained"), "int"),
                "number_of_school_visits": _convert(s.get("Number of School Visits"), "int"),
                "number_of_sessions_conducted": _convert(s.get("Number of Sessions Conducted"), "int"),
                "number_of_modules_completed": _convert(s.get("Number of Modules Completed"), "int"),
                "total_students_intent_rating_per_module": _convert(s.get("Total Students' Intent to Pursue Rating per Module"), "num"),
                "status": _convert(s.get("status"), "text"),
            }
            if row["id"] is not None:
                rep_rows.append(row)

        rep_cols = [
            "id", "ngoid", "month", "number_of_teachers_trained", "number_of_school_visits",
            "number_of_sessions_conducted", "number_of_modules_completed", "total_students_intent_rating_per_module",
            "status"
        ]
        report_inserted = _insert_many(conn, schema, "report", rep_cols, rep_rows, ["id"])
        if report_inserted == 0:
            fallback_ngoid = pre_rows[0]["ngoid"] if pre_rows else "NGO-SAMPLE"
            fallback_report = [{
                "id": 900001,
                "ngoid": fallback_ngoid,
                "month": "Jan-2026",
                "number_of_teachers_trained": 12,
                "number_of_school_visits": 4,
                "number_of_sessions_conducted": 10,
                "number_of_modules_completed": 6,
                "total_students_intent_rating_per_module": 4.2,
                "status": "Submitted",
            }]
            report_inserted = _insert_many(conn, schema, "report", rep_cols, fallback_report, ["id"])
        inserts_summary.append(("report", report_inserted))

        # 6) metrics_base
        met_src = _read_sheet_dicts(wb, "Metrics Base")
        met_rows: List[Dict[str, Any]] = []
        for s in met_src:
            row = {
                "field": _convert(s.get("Field"), "text"),
                "multiplier": _convert(s.get("Mulitplier"), "text"),
                "col_3": _convert(s.get("col_3"), "text"),
                "col_4": _convert(s.get("col_4"), "text"),
                "data_to_be_displayed_on_dashboard": _convert(s.get("Data to be displayed on Dashbaord"), "text"),
            }
            if any(v is not None for v in row.values()):
                met_rows.append(row)

        met_cols = ["field", "multiplier", "col_3", "col_4", "data_to_be_displayed_on_dashboard"]
        inserts_summary.append(("metrics_base", _insert_many(conn, schema, "metrics_base", met_cols, met_rows)))

        # 7) infection
        inf_src = _read_sheet_dicts(wb, "infection")
        inf_rows: List[Dict[str, Any]] = []
        for s in inf_src:
            row = {
                "serial_number": _convert(s.get("serialNumber"), "text"),
                "date": _convert(s.get("date"), "date"),
                "summary": _convert(s.get("summary"), "text"),
                "log": _convert(s.get("log"), "text"),
            }
            if row["serial_number"]:
                inf_rows.append(row)

        inf_cols = ["serial_number", "date", "summary", "log"]
        infection_inserted = _insert_many(conn, schema, "infection", inf_cols, inf_rows, ["serial_number"])
        if infection_inserted == 0:
            fallback_serial = next(iter(laptop_ids), None)
            if fallback_serial:
                fallback_infection = [{
                    "serial_number": fallback_serial,
                    "date": dt.date.today(),
                    "summary": "No major threats found",
                    "log": "Sample antivirus scan log generated by seed script",
                }]
                infection_inserted = _insert_many(conn, schema, "infection", inf_cols, fallback_infection, ["serial_number"])
        inserts_summary.append(("infection", infection_inserted))

        # 8) audit_for_laptops
        with conn.cursor() as cur:
            cur.execute(f"SELECT id FROM {schema}.laptop_labeling")
            laptop_ids = {r[0] for r in cur.fetchall()}

        aud_src = _read_sheet_dicts(wb, "Audit for Laptops")
        aud_rows: List[Dict[str, Any]] = []
        for s in aud_src:
            row = {
                "id": _convert(s.get("ID"), "text"),
                "field": _convert(s.get("Field"), "text"),
                "from_value": _convert(s.get("From"), "text"),
                "to_value": _convert(s.get("To"), "text"),
                "updated_by": _convert(s.get("Updated By"), "text"),
                "updated_on": _convert(s.get("Updated On"), "ts"),
            }
            if row["id"] and row["id"] in laptop_ids:
                aud_rows.append(row)

        aud_cols = ["id", "field", "from_value", "to_value", "updated_by", "updated_on"]
        inserts_summary.append(("audit_for_laptops", _insert_many(conn, schema, "audit_for_laptops", aud_cols, aud_rows)))

        # 9) average_days_count
        avg_src = _read_sheet_dicts(wb, "Average Days Count")
        avg_rows: List[Dict[str, Any]] = []
        for s in avg_src:
            row = {
                "id": _convert(s.get("ID"), "text"),
                "pickup_requested_date": _convert(s.get("Pickup Requested Date"), "date"),
                "distributed_date": _convert(s.get("Distributed Date"), "date"),
                "days_difference": _convert(s.get("Days Difference"), "int"),
                "calculated_on": _convert(s.get("Calculated On"), "date"),
            }
            if row["id"]:
                avg_rows.append(row)

        avg_cols = ["id", "pickup_requested_date", "distributed_date", "days_difference", "calculated_on"]
        inserts_summary.append(("average_days_count", _insert_many(conn, schema, "average_days_count", avg_cols, avg_rows)))

        # 10) laptop_user_map (filtered by existing FK keys)
        with conn.cursor() as cur:
            cur.execute(f"SELECT id FROM {schema}.userdetails")
            user_ids = {r[0] for r in cur.fetchall()}

        map_src = _read_sheet_dicts(wb, "LaptopUserMap")
        map_rows: List[Dict[str, Any]] = []
        for s in map_src:
            lid = _convert(s.get("Laptop ID"), "text")
            uid = _convert(s.get("User ID"), "int")
            row = {
                "laptop_id": lid,
                "user_id": uid,
                "issued_date": _convert(s.get("Issued Date"), "date"),
            }
            if lid and uid is not None and lid in laptop_ids and uid in user_ids:
                map_rows.append(row)

        map_cols = ["laptop_id", "user_id", "issued_date"]
        map_inserted = _insert_many(conn, schema, "laptop_user_map", map_cols, map_rows)
        if map_inserted == 0 and laptop_ids and user_ids:
            fallback_map = [{
                "laptop_id": sorted(laptop_ids)[0],
                "user_id": sorted(user_ids)[0],
                "issued_date": dt.date.today(),
            }]
            map_inserted = _insert_many(conn, schema, "laptop_user_map", map_cols, fallback_map)
        inserts_summary.append(("laptop_user_map", map_inserted))

        # 11) external_registered_ngo (derive from preliminary)
        reg_rows: List[Dict[str, Any]] = []
        for r in pre_rows:
            ngoid = _to_text(r.get("ngoid"))
            doner = _to_text(r.get("doner"))
            if ngoid:
                reg_rows.append({"id": ngoid, "doner": doner})

        reg_cols = ["id", "doner"]
        inserts_summary.append(("external_registered_ngo", _insert_many(conn, schema, "external_registered_ngo", reg_cols, reg_rows, ["id"])))

        # 12) synthetic control data for traceability tables
        with conn.cursor() as cur:
            cur.execute(f"SELECT id, status, working, last_updated_on FROM {schema}.laptop_labeling ORDER BY id LIMIT 1")
            first_laptop = cur.fetchone()

        if first_laptop:
            laptop_id = first_laptop[0]

            # event log
            event_rows = [
                {
                    "laptop_id": laptop_id,
                    "event_type": "STATUS_CHANGE",
                    "field_name": "status",
                    "old_value": "Refurbishment Started",
                    "new_value": first_laptop[1] or "Laptop Refurbished",
                    "actor": "system_seed",
                    "event_time": _utc_now(),
                    "reason": "Sample seed event",
                },
                {
                    "laptop_id": laptop_id,
                    "event_type": "WORKING_UPDATE",
                    "field_name": "working",
                    "old_value": "Not Working",
                    "new_value": first_laptop[2] or "Working",
                    "actor": "system_seed",
                    "event_time": _utc_now(),
                    "reason": "Sample seed event",
                },
            ]
            event_cols = ["laptop_id", "event_type", "field_name", "old_value", "new_value", "actor", "event_time", "reason"]
            inserts_summary.append(("laptop_event_log", _insert_many(conn, schema, "laptop_event_log", event_cols, event_rows)))

            # versions
            with conn.cursor() as cur:
                cur.execute(f"SELECT to_jsonb(t) FROM {schema}.laptop_labeling t WHERE id=%s", (laptop_id,))
                snap = cur.fetchone()[0]
                cur.execute(
                    f"""
                    INSERT INTO {schema}.laptop_versions
                    (laptop_id, version_no, snapshot_json, changed_by, changed_at, change_reason)
                    VALUES (%s, %s, %s::jsonb, %s, %s, %s)
                    ON CONFLICT (laptop_id, version_no) DO NOTHING
                    """,
                    (laptop_id, 1, json.dumps(snap), "system_seed", _utc_now(), "Initial sample snapshot"),
                )
                cur.execute(
                    f"""
                    INSERT INTO {schema}.laptop_versions
                    (laptop_id, version_no, snapshot_json, changed_by, changed_at, change_reason)
                    VALUES (%s, %s, %s::jsonb, %s, %s, %s)
                    ON CONFLICT (laptop_id, version_no) DO NOTHING
                    """,
                    (laptop_id, 2, json.dumps(snap), "system_seed", _utc_now(), "Post-QC sample snapshot"),
                )
            inserts_summary.append(("laptop_versions", 2))

            # stage runs + checklist + qc + feedback
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    INSERT INTO {schema}.laptop_stage_runs
                    (laptop_id, stage_name, run_status, started_by, started_at, completed_by, completed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING stage_run_id
                    """,
                    (laptop_id, "Refurbishment", "completed", "tech_seed", _utc_now(), "tech_seed", _utc_now()),
                )
                stage_refurb = cur.fetchone()[0]

                cur.execute(
                    f"""
                    INSERT INTO {schema}.laptop_stage_runs
                    (laptop_id, stage_name, run_status, started_by, started_at, completed_by, completed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    RETURNING stage_run_id
                    """,
                    (laptop_id, "QC L1", "failed", "qc_seed", _utc_now(), "qc_seed", _utc_now()),
                )
                stage_qc = cur.fetchone()[0]

                checklist_rows = [
                    (stage_refurb, "OS Installed", "yes", "tech_seed", _utc_now(), None),
                    (stage_refurb, "Drivers Installed", "yes", "tech_seed", _utc_now(), None),
                    (stage_refurb, "Battery Captured", "yes", "tech_seed", _utc_now(), None),
                ]
                cur.executemany(
                    f"""
                    INSERT INTO {schema}.laptop_checklist_responses
                    (stage_run_id, checklist_item, response_value, responded_by, responded_at, remark)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    checklist_rows,
                )

                qc_rows = [
                    (laptop_id, stage_qc, "L1", "fail", "Keyboard", "qc_seed", _utc_now(), "3 keys not working"),
                    (laptop_id, stage_qc, "L1", "pass", "Boot", "qc_seed", _utc_now(), "Boot stable"),
                ]
                cur.executemany(
                    f"""
                    INSERT INTO {schema}.qc_checks
                    (laptop_id, stage_run_id, qc_layer, qc_result, defect_type, checked_by, checked_at, remark)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    qc_rows,
                )

                cur.execute(
                    f"""
                    INSERT INTO {schema}.issue_feedback
                    (laptop_id, issue_source, issue_category, severity, reported_by, reported_at, resolution_action, resolved_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        laptop_id,
                        "ngo",
                        "keyboard",
                        "medium",
                        "ops_seed",
                        _utc_now(),
                        "Keyboard replaced",
                        _utc_now(),
                    ),
                )
            inserts_summary.append(("laptop_stage_runs", 2))
            inserts_summary.append(("laptop_checklist_responses", 3))
            inserts_summary.append(("qc_checks", 2))
            inserts_summary.append(("issue_feedback", 1))

        conn.commit()

        print("Seeding completed successfully")
        for name, count in inserts_summary:
            print(f"- {name}: {count} rows attempted")

    except Exception:
        conn.rollback()
        raise
    finally:
        close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Seed sheet-first schema tables from Dev Laptop Data.xlsx")
    parser.add_argument(
        "--database-url",
        default=os.getenv("DATABASE_URL", ""),
        help="Postgres connection URL. Defaults to env DATABASE_URL",
    )
    parser.add_argument(
        "--schema",
        default=os.getenv("DB_SCHEMA", "sama_ops"),
        help="Target schema name. Default: sama_ops",
    )
    parser.add_argument(
        "--workbook",
        default="Dev Laptop Data.xlsx",
        help="Path to workbook. Default: Dev Laptop Data.xlsx",
    )
    parser.add_argument(
        "--truncate",
        action="store_true",
        help="Truncate managed tables before seeding.",
    )

    args = parser.parse_args()

    if not args.database_url:
        print("Error: --database-url is required (or set DATABASE_URL).", file=sys.stderr)
        return 2

    try:
        seed(args.database_url, args.schema, Path(args.workbook), args.truncate)
        return 0
    except Exception as exc:
        print(f"Seeding failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
