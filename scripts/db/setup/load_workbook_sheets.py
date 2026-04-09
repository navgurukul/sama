#!/usr/bin/env python3
"""Load every sheet from an Excel workbook into PostgreSQL tables.

Table naming:
- <table_prefix>_<sheet_name_sanitized>

Column naming:
- Sanitized from header row
- Empty headers become col_<n>
- Duplicate headers receive numeric suffixes

Examples:
  python scripts/db/setup/load_workbook_sheets.py --database-url "postgresql://..." --schema sama_ops --workbook "User-Roles.xlsx"
  python scripts/db/setup/load_workbook_sheets.py --database-url "postgresql://..." --schema sama_ops --workbook "User-Roles.xlsx" --table-prefix user_roles --reset
"""

from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple


def _connect(database_url: str):
    try:
        import psycopg  # type: ignore

        conn = psycopg.connect(database_url)

        def close() -> None:
            conn.close()

        return conn, close
    except Exception as exc:
        raise RuntimeError("psycopg is required. Install with: pip install psycopg[binary]") from exc


def _load_workbook(path: Path):
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from exc

    return load_workbook(path, data_only=True)


def _sanitize_identifier(text: str, fallback: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9_]+", "_", text.strip().lower())
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        s = fallback
    if s[0].isdigit():
        s = f"n_{s}"
    return s


def _ensure_valid_identifier(name: str, label: str) -> str:
    if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
        raise ValueError(f"Invalid {label}: {name!r}")
    return name


def _row_is_empty(row: Sequence[Any]) -> bool:
    return all(v is None or str(v).strip() == "" for v in row)


def _cell_to_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    return text if text != "" else None


def _dedupe_headers(headers: Sequence[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    for i, h in enumerate(headers, start=1):
        base = _sanitize_identifier(h, f"col_{i}")
        count = seen.get(base, 0) + 1
        seen[base] = count
        out.append(base if count == 1 else f"{base}_{count}")
    return out


def _iter_sheet_rows(ws) -> Iterable[Tuple[List[str], List[Tuple[Any, ...]]]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    last_non_empty = -1
    for row in rows:
        for idx, value in enumerate(row):
            if value is not None and str(value).strip() != "":
                last_non_empty = max(last_non_empty, idx)

    if last_non_empty < 0:
        return

    width = last_non_empty + 1
    raw_headers = ["" if h is None else str(h) for h in rows[0][:width]]
    headers = _dedupe_headers(raw_headers)
    data_rows: List[Tuple[Any, ...]] = []

    for row in rows[1:]:
        if _row_is_empty(row):
            continue
        normalized = []
        for idx in range(len(headers)):
            value = row[idx] if idx < len(row) else None
            normalized.append(_cell_to_text(value))
        data_rows.append(tuple(normalized))

    yield headers, data_rows


def _create_table(conn, schema: str, table: str, columns: List[str], reset: bool, truncate: bool) -> None:
    with conn.cursor() as cur:
        if reset:
            cur.execute(f"DROP TABLE IF EXISTS {schema}.{table}")

        col_sql = ",\n  ".join(f"{c} text" for c in columns)
        cur.execute(
            f"""
            CREATE TABLE IF NOT EXISTS {schema}.{table} (
              {col_sql}
            )
            """
        )

        if truncate:
            cur.execute(f"TRUNCATE TABLE {schema}.{table}")


def _insert_rows(conn, schema: str, table: str, columns: List[str], rows: List[Tuple[Any, ...]]) -> int:
    if not rows:
        return 0
    placeholders = ", ".join(["%s"] * len(columns))
    col_sql = ", ".join(columns)
    with conn.cursor() as cur:
        cur.executemany(
            f"INSERT INTO {schema}.{table} ({col_sql}) VALUES ({placeholders})",
            rows,
        )
    return len(rows)


def load_workbook_to_tables(database_url: str, schema: str, workbook: Path, table_prefix: str, reset: bool, truncate: bool) -> None:
    schema = _ensure_valid_identifier(_sanitize_identifier(schema, "sama_ops"), "schema")
    table_prefix = _ensure_valid_identifier(_sanitize_identifier(table_prefix, "workbook"), "table prefix")

    if not workbook.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook}")

    conn, close = _connect(database_url)
    wb = _load_workbook(workbook)

    try:
        with conn.cursor() as cur:
            cur.execute(f"CREATE SCHEMA IF NOT EXISTS {schema}")

        created: List[Tuple[str, int, int]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for columns, data_rows in _iter_sheet_rows(ws):
                table_name = _sanitize_identifier(f"{table_prefix}_{sheet_name}", f"{table_prefix}_sheet")
                table_name = _ensure_valid_identifier(table_name, "table name")
                _create_table(conn, schema, table_name, columns, reset=reset, truncate=truncate)
                inserted = _insert_rows(conn, schema, table_name, columns, data_rows)
                created.append((table_name, len(columns), inserted))

        conn.commit()

        print(f"Workbook loaded successfully: {workbook.name}")
        for table_name, col_count, row_count in created:
            print(f"- {schema}.{table_name}: {col_count} columns, {row_count} rows inserted")

    except Exception:
        conn.rollback()
        raise
    finally:
        close()


def main() -> int:
    parser = argparse.ArgumentParser(description="Load workbook sheets to PostgreSQL tables.")
    parser.add_argument("--database-url", default=os.getenv("DATABASE_URL", ""), help="Postgres connection URL")
    parser.add_argument("--schema", default=os.getenv("DB_SCHEMA", "sama_ops"), help="Target schema")
    parser.add_argument("--workbook", required=True, help="Workbook path")
    parser.add_argument("--table-prefix", default="", help="Table name prefix. Default: workbook filename stem")
    parser.add_argument("--reset", action="store_true", help="Drop/recreate target workbook tables")
    parser.add_argument("--truncate", action="store_true", help="Truncate target workbook tables before insert")

    args = parser.parse_args()

    if not args.database_url:
        print("Error: --database-url is required (or set DATABASE_URL)", file=sys.stderr)
        return 2

    workbook = Path(args.workbook)
    table_prefix = args.table_prefix or workbook.stem

    try:
        load_workbook_to_tables(
            database_url=args.database_url,
            schema=args.schema,
            workbook=workbook,
            table_prefix=table_prefix,
            reset=args.reset,
            truncate=args.truncate,
        )
        return 0
    except Exception as exc:
        print(f"Workbook load failed: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
