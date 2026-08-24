from openpyxl import load_workbook
import sys

path = sys.argv[1]
wb = load_workbook(path, data_only=True)
print('SHEETS:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0] if rows else []
    norm_headers = [str(x).strip() if x is not None else '' for x in headers]
    print(f'--- {sheet} ---')
    print('HEADERS:', norm_headers)
    print('ROW_COUNT_EXCL_HEADER:', max(0, len(rows) - 1))
